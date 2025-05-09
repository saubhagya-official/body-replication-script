import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

# File paths
abbr_file = './resources/abbreviations.xlsx'
mapping_file = './resources/mappings.xlsx'
main_file = './resources/base.xlsx'
output_file = 'output.xlsx'

# Sheet names
abbr_sheet = 'Sheet1'
mapping_sheet = 'Sheet1'
main_sheet = 'ProductModelPickOptions'

# Variations
s3_variation = 's3'  # Target variation
s2_variation = 's2'  # Source variation

# Validate files and sheets
for file, name in [(abbr_file, 'abbreviations'), (mapping_file, 'mapping'), (main_file, 'main')]:
    try:
        excel = pd.ExcelFile(file)
        print(f"Available worksheets in '{file}' ({name}): {excel.sheet_names}")
    except FileNotFoundError:
        raise FileNotFoundError(f"{name.capitalize()} file '{file}' not found.")

if abbr_sheet not in pd.ExcelFile(abbr_file).sheet_names:
    raise ValueError(f"Sheet '{abbr_sheet}' not found in '{abbr_file}'.")
if mapping_sheet not in pd.ExcelFile(mapping_file).sheet_names:
    raise ValueError(f"Sheet '{mapping_sheet}' not found in '{mapping_file}'.")
if main_sheet not in pd.ExcelFile(main_file).sheet_names:
    raise ValueError(f"Sheet '{main_sheet}' not found in '{main_file}'.")

# Load DataFrames
try:
    abbr_df = pd.read_excel(abbr_file, sheet_name=abbr_sheet)
    if len(abbr_df.columns) >= 3:
        abbr_df.columns = ['code', 'variation', 'description'] + list(abbr_df.columns[3:])
    else:
        raise ValueError(f"Expected at least 3 columns in {abbr_file}, found {len(abbr_df.columns)}")
    mapping_df = pd.read_excel(mapping_file, sheet_name=mapping_sheet)
    main_df = pd.read_excel(main_file, sheet_name=main_sheet)
except Exception as e:
    raise Exception(f"Error loading sheets: {str(e)}")

# Clean abbr_df: Drop empty rows
abbr_df = abbr_df.dropna(how='all')

# Debug: Check loaded data
print("\nFirst few rows of abbr_df:")
print(abbr_df.head())
print("\nmapping_df columns:", mapping_df.columns.tolist())
print("\nFirst few rows of main_df:")
print(main_df.head())

# Validate required columns in mapping_df
required_mapping_cols = ['Description', 'ProductModelID']
if not all(col in mapping_df.columns for col in required_mapping_cols):
    missing = [col for col in required_mapping_cols if col not in mapping_df.columns]
    raise ValueError(f"Missing columns in {mapping_file}: {missing}")

# Ensure 'variation' is a string
abbr_df['variation'] = abbr_df['variation'].astype(str)

# Filter for s3 and s2 variations
s3_df = abbr_df[abbr_df['variation'].str.contains(s3_variation, case=False, na=False)]
if s3_df.empty:
    print(f"Warning: No rows found with '{s3_variation}' in variation column.")
    print("Available variations:", abbr_df['variation'].unique())
    raise ValueError(f"No rows found with '{s3_variation}' in variation column.")

s2_df = abbr_df[abbr_df['variation'].str.contains(s2_variation, case=False, na=False)]
if s2_df.empty:
    print(f"Warning: No rows found with '{s2_variation}' in variation column.")
    print("Available variations:", abbr_df['variation'].unique())
    raise ValueError(f"No rows found with '{s2_variation}' in variation column.")

# Check for duplicate codes
duplicate_s3_codes = s3_df[s3_df['code'].duplicated()]['code'].unique()
if len(duplicate_s3_codes) > 0:
    print(f"Warning: Duplicate {s3_variation} entries for codes: {duplicate_s3_codes}")
duplicate_s2_codes = s2_df[s2_df['code'].duplicated()]['code'].unique()
if len(duplicate_s2_codes) > 0:
    print(f"Warning: Duplicate {s2_variation} entries for codes: {duplicate_s2_codes}")

# Map codes to descriptions
code_to_s3_desc = s3_df.groupby('code').first()['description'].to_dict()
code_to_s2_desc = s2_df.groupby('code').first()['description'].to_dict()

# Map descriptions to ProductModelID
desc_to_number = mapping_df.set_index('ModelNumber')['ProductModelID'].astype(str).to_dict()

# Create replace pairs: s2 ProductModelID to s3 ProductModelID
replace_pairs = []
missing_count = 0
s2_codes = s2_df['code'].unique()
for code in s2_codes:
    # Get s2 description and ProductModelID
    s2_desc = code_to_s2_desc.get(code)
    if s2_desc and s2_desc in desc_to_number:
        s2_number = desc_to_number[s2_desc]
        # Get s3 description and ProductModelID
        s3_desc = code_to_s3_desc.get(code)
        if s3_desc and s3_desc in desc_to_number:
            s3_number = desc_to_number[s3_desc]
            replace_pairs.append([s2_number, s3_number])
        else:
            print(f"Warning: {s3_variation} description '{s3_desc}' for code '{code}' not found in mappings.")
            missing_count += 1
    else:
        print(f"Warning: {s2_variation} description '{s2_desc}' for code '{code}' not found in mappings or code missing in s3.")
        missing_count += 1
if missing_count > 0:
    print(f"Total missing mappings: {missing_count}")
if not replace_pairs:
    raise ValueError(f"No replace pairs created for {s2_variation} to {s3_variation}. Check mappings or descriptions.")

print("\nReplace pairs (s2 ProductModelID to s3 ProductModelID):", replace_pairs)

# Check for s2 ProductModelID in key columns
key_columns = [col for col in main_df.columns if col.startswith('Key')]
s2_numbers = [pair[0] for pair in replace_pairs]
found_numbers = False
for col in key_columns:
    if main_df[col].astype(str).str.contains('|'.join(s2_numbers), case=False, na=False).any():
        found_numbers = True
        print(f"s2 ProductModelID found in column: {col}")
if not found_numbers:
    print(f"Warning: None of the s2 ProductModelID {s2_numbers} found in key columns {key_columns} of {main_sheet}.")
    print("Checking all sheets...")
    excel = pd.ExcelFile(main_file)
    for sheet in excel.sheet_names:
        df = pd.read_excel(main_file, sheet_name=sheet)
        if df.astype(str).apply(lambda x: x.str.contains('|'.join(s2_numbers), case=False, na=False)).sum().sum() > 0:
            print(f"s2 ProductModelID found in sheet: {sheet}")

# Create a copy of the main DataFrame
main_df_copy = main_df.copy()

# Track replacements
replacements_log = []

# Perform find-and-replace in key columns
for old_value, new_value in replace_pairs:
    for col in key_columns:
        mask = main_df_copy[col].astype(str) == old_value
        if mask.any():
            for row_idx, is_match in mask.items():
                if is_match:
                    replacements_log.append({
                        'row': row_idx + 2,
                        'column': main_df_copy.columns.get_loc(col) + 1,
                        'column_name': col,
                        'old_value': old_value,
                        'new_value': new_value
                    })
            main_df_copy[col] = main_df_copy[col].astype(str).replace(old_value, new_value)

# Load the original workbook
try:
    workbook = openpyxl.load_workbook(main_file)
except Exception as e:
    raise Exception(f"Failed to load {main_file}: {str(e)}")
original_sheet = workbook[main_sheet]
new_sheet = workbook.create_sheet(main_sheet)

# Copy cell formatting
for row in original_sheet:
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row, column=cell.column)
        new_cell._style = copy(cell._style)

# Remove the original sheet
workbook.remove(original_sheet)

# Write modified data
data = [main_df_copy.columns.tolist()] + main_df_copy.values.tolist()
for row_idx, row in enumerate(data, start=1):
    for col_idx, value in enumerate(row, start=1):
        new_sheet.cell(row=row_idx, column=col_idx, value=value)

# Highlight replaced cells
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
for replacement in replacements_log:
    cell = new_sheet.cell(row=replacement['row'], column=replacement['column'])
    cell.fill = yellow_fill

# Save the workbook
try:
    workbook.save(output_file)
except Exception as e:
    raise Exception(f"Failed to save {output_file}: {str(e)}")

# Log replacements
print(f"\nModified spreadsheet saved as '{output_file}'")
if replacements_log:
    print(f"\nReplacements performed ({len(replacements_log)} total):")
    for r in replacements_log:
        print(f"Row {r['row']}, Column {r['column']} ({r['column_name']}): "
              f"{r['old_value']} → {r['new_value']}")
else:
    print("No replacements were made.")